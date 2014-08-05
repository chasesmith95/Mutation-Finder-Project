''''

author=Chase Smith
date= July 23, 2014



'''


from Bio import SeqIO
from Bio import Entrez
from Bio.Seq import Seq 
from Bio.Seq import MutableSeq
from Bio.Alphabet import IUPAC
from openpyxl import load_workbook
from openpyxl import *
import urllib.request
import time
import csv
from scipy import *
from pyDOE.doe_factorial import *
import os
import glob
from mechanicalsoup import * 
#os.sys.path.append(mechanize)

#from mechanize.mechanize import _mechanize as mechanize


Entrez.email ="chase.smth@gmail.com"




class Fasta(object):	
	
	def __init__(self, input_file, input_sheet, input_list, output_file):
		self.input_file = str(input_file)
		self.input_sheet=str(input_sheet)
		self.input_list=input_list
		self.output_file=str(output_file)
		self.accessionId_col=input_list[0]
		self.accessionId=[]
		self.mutIndex_col=input_list[1]
		self.mutIndex=[]
		self.aChange_col=input_list[2]
		self.aChange=[]
		self.numDict={'A': 0, 'B': 1, 'C':2, 'D': 3, 'E': 4, 'F' : 5, 'G':6 ,'H' :7, 'I' :8, 'J' :9, 'K' : 10, 'L' : 11}
		
		self.mutatedProtein=[]
		self.proteinId=[]
		self.fastaHeader=[]
		
	
		
	def processMutatedProteinFasta(self):
		Entrez.email ="chase.smth@gmail.com"
		print("reading input")
		self.readInputAccessionId(self)
		print("done")
		i=0
		while( i< len(self.accessionId)):
			if(self.mutIndex[i]!='-' and self.aChange[i]!='-'):
				print('Looking up the Protein')
				self.lookUpProteinFasta(i)
				print(self.proteinId[(i-1)])
				print(len(self.proteinId))
				self.getFastaHeader(i)
			elif(self.mutIndex[i]=='-' or self.aChange[i]=='-'):
				self.mutatedProtein.append('-')
				self.proteinId.append('-')
				self.fastaHeader.append('-')	
			i=i+1
		print("Writing to a file")
		self.writeFasta(self)
		return
	def lookUpProteinFasta(self, k):
		i=k
		muteIndex=int(float(self.mutIndex[i]))
		handle = Entrez.esearch(db="nucleotide", term=str(self.accessionId[i]))
		record = Entrez.read(handle)
		temp=record["IdList"]
		#download the info you want
		download = Entrez.efetch(db="nucleotide", id=temp, rettype="gb", retmode="xml")
		sequenceData = Entrez.read(download)
		#print(sequenceData)
		#turn a selection of this info into an str which is then isolated to find protein
		largeData=sequenceData[0]["GBSeq_feature-table"]
		l=0
		while l<len(largeData):
			j=0
			smallData=sequenceData[0]["GBSeq_feature-table"][l]['GBFeature_quals']
			while j<len(smallData): 
				if(sequenceData[0]["GBSeq_feature-table"][l]['GBFeature_quals'][j]['GBQualifier_name']=='protein_id'):
					tempId=str(sequenceData[0]["GBSeq_feature-table"][l]['GBFeature_quals'][j]['GBQualifier_value'])
					print(tempId)
					self.proteinId.append(tempId)
				if(sequenceData[0]["GBSeq_feature-table"][l]['GBFeature_quals'][j]['GBQualifier_name']=='translation'):
					tempProtein= sequenceData[0]["GBSeq_feature-table"][l]['GBFeature_quals'][j]['GBQualifier_value']
					mutProtein= str(tempProtein[0:muteIndex-1] + self.aChange[i][-1] + tempProtein[muteIndex:])
					self.mutatedProtein.append(mutProtein)
					print(mutProtein)
				j=j+1
			l=l+1
		return
	def getFastaHeader(self, j):
		k=j
		i=j-1
		handle = Entrez.esearch(db="protein", term=str(self.proteinId[i]))
		record = Entrez.read(handle)
		temp=record["IdList"]
			#download the info you want
		download = Entrez.efetch(db="protein", id=temp, rettype="fasta", retmode="xml")
		sequenceData = Entrez.read(download)
		print(sequenceData)
		proteinDefinition=sequenceData[0]['TSeq_defline']
		print(sequenceData[0]['TSeq_defline'])
		fastaGeneId= sequenceData[0]['TSeq_gi']
		fastaHeaderStr='>gi|'+ fastaGeneId + '|ref|' + self.proteinId[i] + '|' + proteinDefinition +'|' + self.aChange[k][0] + str(int(self.mutIndex[k])) + self.aChange[k][-1]
		self.fastaHeader.append(fastaHeaderStr)
		return 
	def writeFasta(self):
		file = open(self.output_file, 'w')
		i=0
		while(i<len(self.proteinId)):
			temp=''
			j=80
			if(self.mutatedProtein[i]=='-'):
				i=i
			elif(self.mutatedProtein[i]!='-'):
				while( len(self.mutatedProtein[i])-j>80):
					temp= temp + '\n' + str(self.mutatedProtein[i][(j-80):j]) 
					j=j+80
				temp=temp + '\n' + str(self.mutatedProtein[i])[(j):]
				fastaStr=(str(self.fastaHeader[i]) + temp)
				file.write(fastaStr +'\n' +'\n')
			i=i+1
		file.close()
		return
	def readInputAccessionId(self):   
		wb = load_workbook(filename = self.input_file, use_iterators = True)
		ws = wb.get_sheet_by_name(name = self.input_sheet) # ws is now an IterableWorksheet
		
		for row in ws.iter_rows(): # it brings a new method: iter_rows()
			i=0
			for cell in row:
				if(i==(self.numDict(self.accessionId_col)) and (cell.value)!='None'):
					self.accessionId.append(cell.value)
				if(i== (self.numDict(self.mutIndex_col))  and (cell.value)!='None'):
					self.mutIndex.append(cell.value)
				if(i== (self.numDict(self.aChange_col))  and (cell.value)!='None'):
					self.aChange.append(cell.value)
				i=i+1

	def readInputEnsembl(self):	
			try:
				wb = load_workbook(filename = self.input_file, use_iterators = True)
				ws = wb.get_sheet_by_name(name = self.input_sheet)
				for row in ws.iter_rows(): # it brings a new method: iter_rows()
					i=0
					for cell in row:
						if(i==(self.numDict(self.Ensembl_col))and (cell.value)!='None'):
							fullEnsembl=str(cell.value)
							fullEnsembl=list(fullEnsembl.replace(':p.', ','))
							aChangeStart=self.aminoAcidDict[fullEnsembl[1][0:3]]
							aChangeEnd=self.aminoAcidDict[fullEnsembl[1][6:]]
							self.aChange.append(aChangeStart + '/' + aChangeEnd )
							self.mutIndex.append(fullEnsembl[1][3:])
							self.EnsemblId.append(fullEnsembl[0])
						elif((cell.value)=='None'):
							self.aChange.append('-')
							self.mutIndex.append('-')
							self.EnsemblId.append('-')
						i=i+1
					return
			except:
				with open(self.input_file, 'r') as file:				
					wb = csv.reader(file, delimiter='\t', quoting=csv.QUOTE_NONE)
					for row in wb:
						line=str(row[27])
						print(line)
						if(line=='Amino acid change in longest transcript' or line=='' or line=='-' or line=='None'):
							line='-'
						else:
							fullEnsembl=(line.replace(':p.', ','))
							fullEnsembl=list(fullEnsembl.split(','))
							self.EnsemblId.append(str(fullEnsembl[0]))
							print(str(fullEnsembl[0]))	
							i=0
							index=3
							count=0
							while i< len(fullEnsembl[1]):
								if(fullEnsembl[1][i].isdigit()==True):
									index=i
									count=count+1
								i=i+1
							print(fullEnsembl[1][index-count+1:index+1])
							print(fullEnsembl[0])	
							self.mutIndex.append(fullEnsembl[1][index-count+1:index+1])		
							aChangeStart=self.aminoAcidDict[fullEnsembl[1][0:index-count+1]]
							aChangeEnd=self.aminoAcidDict[fullEnsembl[1][index+1:]]
							self.aChange.append(aChangeStart + '/' + aChangeEnd )
							print(aChangeStart)
							
							#listIds=line.replace()
							print(line)
					return
	




class Excel(object):
	
	
	def readInputAccessionId(self):   
		try:
				wb = load_workbook(filename = self.input_file, use_iterators = True)
				ws = wb.get_sheet_by_name(name = self.input_sheet) # ws is now an IterableWorksheet
				for row in ws.iter_rows(): # it brings a new method: iter_rows()
					i=0
					for cell in row:
						if(i==(self.numDict[self.accessionId_col])and (cell.value)!='None'):
							self.accessionId.append(cell.value)
						if(i== (self.numDict[self.mutIndex_col]) and (cell.value)!='None'):
							self.mutIndex.append(cell.value)
						if(i==(self.numDict[self.aChange_col]) and (cell.value)!='None'):
							self.aChange.append(cell.value)
						i=i+1
		except:
				with open(self.input_file, 'r') as file:				
					wb = csv.reader(file, delimiter='\t', quoting=csv.QUOTE_NONE)
					for row in wb:
						self.accessionId.append(row[self.numDict[self.accessionId_col]])
						self.mutIndex.append(row[self.numDict[self.mutIndex_col]])
						self.aChange.append((self.numDict[self.aChange_col]))
					return
	
	def readInputEnsembl(self):	
			try:
				wb = load_workbook(filename = (self.input_file), use_iterators = True)
				print("Load Workbook Successful")
				#print(str(self.input_sheet))
				ws = wb.get_sheet_by_name(name = self.input_sheet)
				#print(str(self.input_sheet))
			
				for row in (ws.iter_rows()): # it brings a new method: iter_rows()
					i=0
					#print(self.numDict[str(self.Ensembl_col)])
					for cell in row:
						#print(cell.value)
						
						
						if(i==(int(self.numDict[str(self.Ensembl_col)])) and (cell.value)!='None' and (cell.value)!=''):
							#print("Here")
							self.fullEnsembl.append((cell.value))
						elif(((cell.value)=='None' or (cell.value)=='' ) and i==(int(self.numDict[str(self.Ensembl_col)]))):
							self.fullEnsembl.append('-')
						i=i+1
				#print("FULL ENsembl" + self.fullEnsembl)
				i=0
				while(i<len(self.fullEnsembl)):
					try:
						if((self.fullEnsembl[i]).find('_')<0 and (self.fullEnsembl[i]).find('fs')<0 and  ((self.fullEnsembl)[i].find('de'))<0 and (self.fullEnsembl)[i].find('ENSP')>=0):
							fullEn=str(((self.fullEnsembl)[i]).replace(':p.', ','))
							fullEn=list(fullEn.split(','))
							
							k=0
							index=3
							count=0
							while k< len(fullEn[1]):
								if(fullEn[1][k].isdigit()==True):
									index=k
									count=count+1
								k=k+1
							#print(fullEn[1][index-count+1:index+1])
							
							#print(fullEn)
							aChangeStart=self.aminoAcidDict[fullEn[1][0:3]]
							aChangeEnd=self.aminoAcidDict[fullEn[1][index+1:]]
							aCH=aChangeStart + '/' + aChangeEnd 
							self.aChange.append(aCH)
							self.mutIndex.append(fullEn[1][index-count+1:index+1])
							self.EnsemblId.append(fullEn[0])
						elif((self.fullEnsembl[i]).find('_')>=0 or ((self.fullEnsembl)[i].find('de'))>=0 or ((self.fullEnsembl)[i].find('de'))>=0 or (self.fullEnsembl)[i]=='-'):
							self.aChange.append('-')
							self.mutIndex.append('-')
							self.EnsemblId.append('-')
						i=i+1
					except:
						self.aChange.append('-')
						self.mutIndex.append('-')
						self.EnsemblId.append('-')
						i=i+1
				print(self.aChange)
				print(self.mutIndex)
				print(self.EnsemblId)
				return
			except:
				with open(self.input_file, 'r') as file:				
					wb = csv.reader(file, delimiter='\t', quoting=csv.QUOTE_NONE)
					for row in wb:
						line=str(row[int(self.numDict(self.Ensembl_col))])
						#print(line)
						if(line=='Amino acid change in longest transcript' or line=='' or line=='-' or line=='None'):
							line='-'
						else:
							fullEnsembl=(line.replace(':p.', ','))
							fullEnsembl=list(fullEnsembl.split(','))
							self.EnsemblId.append(str(fullEnsembl[0]))
							print(str(fullEnsembl[0]))	
							i=0
							index=3
							count=0
							while i< len(fullEnsembl[1]):
								if(fullEnsembl[1][i].isdigit()==True):
									index=i
									count=count+1
								i=i+1
							print(fullEnsembl[1][index-count+1:index+1])
							print(fullEnsembl[0])	
							self.mutIndex.append(fullEnsembl[1][index-count+1:index+1])		
							aChangeStart=self.aminoAcidDict[fullEnsembl[1][0:index-count+1]]
							aChangeEnd=self.aminoAcidDict[fullEnsembl[1][index+1:]]
							self.aChange.append(aChangeStart + '/' + aChangeEnd )
							#print(aChangeStart)
							
							#listIds=line.replace()
							#print(line)
					return
	
	def runExcel(self):
		print(self.output_file)
		print(self.output_file.find('txt'))
		start=time.time()
		Entrez.email ="chase.smth@gmail.com"
		self.processMutationLists()
		print("Writing to file.....")
		try:
			if(int(self.output_file.find('xlsx'))>=0):
				print("writing to xlsx")
				self.writeToXLSX(self, 0)
			elif(int(self.ouput_file.find('txt'))>=0 ):
				print("writing to csv")
				self.csvWrite(0)
			else:
				print("Unknown")
				self.csvWrite(0)
				end=time.time()-start
				print(end)
		except:
			print("writing to csv")
			self.csvWrite(0)
			end=time.time()-start
			print((end))
	def lookUpEnsemblProtein(self, k):
		i=k
		#print(self.EnsemblId[i])
		Query="http://www.uniprot.org/uniprot/?query=" + str((self.EnsemblId)[i]) + "&sort=score"
		response = urllib.request.Request(Query)
		html = (urllib.request.urlopen(response))
		readText=html.read()
		readText=readText.decode()
		html.close()
		start=readText.find('/span></th></tr><tr><td class="checkboxColumn"><input onclick="addOrAppendCart(')
		end= readText.find(')" class="cart-item" id="checkbox_')	
				#1 first need to isolate the table 
		readText=readText[start:end]
		start=readText.find("('")
		readText=readText[start+2:-1]
		#now must query the entry id with the uniprot database
		queryUni= "http://www.uniprot.org/uniprot/"
		newResponse= (urllib.request.Request(queryUni + readText))
		newHtml = (urllib.request.urlopen(newResponse))
		readText=newHtml.read().decode()
		newHtml.close()
		start=readText.find('<pre class="sequence">')
		readText=readText[start:]
		end= readText.find('</pre>')
		readText=readText[:end]
		##print(readText)
		readText=list(readText.splitlines())
		#print(readText)
		finalString=''
		for lines in readText:
			if(str(lines)=='' or str(lines).find('>')>=0):
				lines=lines
			elif(str(lines[0]).isupper()):
				finalString= finalString + str(lines)
		finalString=finalString.replace(' ', '')
		#print(finalString[0:int(len(finalString)/2)])
		#print(finalString)
		self.proteinSeq.append(finalString)
		return
		

	def lookUpProtein(self,k):
		i=k
		handle = Entrez.esearch(db="nucleotide", term=str(self.accessionId[i]))
		record = Entrez.read(handle)
		temp=record["IdList"]
		#download the info you want
		download = Entrez.efetch(db="nucleotide", id=temp, rettype="gb", retmode="xml")
		sequenceData = Entrez.read(download)
		#print(sequenceData)
		#turn a selection of this info into an str which is then isolated to find protein
		largeData=sequenceData[0]["GBSeq_feature-table"]
		l=0
		while l<len(largeData):
			j=0
			smallData=sequenceData[0]["GBSeq_feature-table"][l]['GBFeature_quals']
			while j<len(smallData): 
				if(sequenceData[0]["GBSeq_feature-table"][l]['GBFeature_quals'][j]['GBQualifier_name']=='translation'):
					tempProtein= sequenceData[0]["GBSeq_feature-table"][l]['GBFeature_quals'][j]['GBQualifier_value']
					self.proteinSeq.append(tempProtein)
					#print(tempProtein)
				j=j+1
			l=l+1
		return
	
	def getMutSeq(self):
		print("start...")
		
		merLength=int(self.amerLength)
		#print(merLength)
		amerLength=int(merLength-1)
		i=0
		while(i<len(self.proteinSeq)):
			try:
				if((self.proteinSeq)[i]=='-'):
					self.mutSeq.append('-')
				if(self.proteinSeq[i]!='-'):
					start=int(float((self.mutIndex)[i]))-merLength
					end=int(float((self.mutIndex)[i])+amerLength)
					temp=self.aChange[i]
					if((int(float((self.mutIndex)[i])))<merLength):
						start=0
					if((len(self.proteinSeq[i])-int(float(self.mutIndex[i])))<amerLength):
						end=len((self.proteinSeq)[i])
					if(temp.find('*')<0 and temp.find('?')<0):
						temp=temp[-1]
					if(temp.find('*')>=0 or temp.find('?')>=0):
						temp=''
						end=int(float(self.mutIndex[i]))+1
					
					sequenceProtein = str((self.proteinSeq)[i])
					smallRegSeq=str(sequenceProtein[start:(int(float((self.mutIndex)[i]))-1)] + temp + sequenceProtein[int(float((self.mutIndex)[i])): end])
					self.mutSeq.append(smallRegSeq)
					#print(self.EnsemblId[i] + "  " + smallRegSeq)
				i=i+1
			except:
				i=i+1	
	def getRegSeq(self):
		print("start...")
		merLength=int(self.amerLength)
		amerLength=int(merLength-1)
		i=0
		if(self.Ensembl):
			while(i<len(self.proteinSeq)):
				try:
					if((self.proteinSeq)[i]=='-' or (self.proteinSeq)[i]=='' or (self.EnsemblId)[i]=='-'):
						self.regSeq.append('-')
					if(self.proteinSeq[i]!='-' and self.proteinSeq[i]!='' and self.EnsemblId[i]!='-'):
						start=(int(float(self.mutIndex[i]))-int(amerLength))
						end=(int(float(self.mutIndex[i])+int(merLength)))
						if((int(float(self.mutIndex[i])))<=int(amerLength)):
							start=0
						if((len((self.proteinSeq)[i])-int(float((self.mutIndex)[i])))<=int(merLength)):
							end=len((self.proteinSeq)[i])
						sequenceProtein = str((self.proteinSeq)[i])
						smallRgSeq=str(sequenceProtein[start:end])
						self.regSeq.append(smallRgSeq)
						#print((self.EnsemblId)[i] + " " + smallRgSeq)
					i=i+1
				except:
					i=i+1
		elif(self.Ensembl==False):
			while(i<len(self.proteinSeq)):
				try:
					if((self.proteinSeq)[i]=='-' or (self.proteinSeq)[i]=='' or (self.accessionId)[i]=='-'):
						self.regSeq.append('-')
					if(self.proteinSeq[i]!='-' and self.proteinSeq[i]!='' and self.accessionId[i]!='-'):
						start=(int(float(self.mutIndex[i]))-int(amerLength))
						end=(int(float(self.mutIndex[i])+int(merLength)))
						if((int(float(self.mutIndex[i])))<=int(amerLength)):
							start=0
						if((len((self.proteinSeq)[i])-int(float((self.mutIndex)[i])))<=int(merLength)):
							end=len((self.proteinSeq)[i])
						sequenceProtein = str((self.proteinSeq)[i])
						smallRgSeq=str(sequenceProtein[start:end])
						self.regSeq.append(smallRgSeq)
						#print((self.EnsemblId)[i] + " " + smallRgSeq)
					i=i+1
				except:
					i=i+1
		print("done")
	
	
	def csvWrite(self, k):
		if(self.Ensembl):
			k=k
			csvoutfile=open((self.output_file), 'w',newline="") 	
			filewriter=csv.writer(csvoutfile, delimiter='\t' )
			filewriter.writerow(np.array(self.headerValues))
			j=0
			i=0
			while j< len(self.EnsemblId):
				while (i< len(list(self.mutSYSeq)[i]) and i<len(list(self.regSYSeq)[i])):
					lines=[str(list(self.EnsemblId)[j]), str(list(self.mutIndex)[j]), str(list(self.aChange)[j]), str(list(self.mutSeq)[j]),
					str(list(self.regSeq)[j]),
					str(list(self.mutSYSeq)[j][i]),
					str(list(self.regSYSeq)[j][i]),
					str(list(self.mutSYStrength)[j][i]),
					str(list(self.regSYStrength)[j][i]),
					str(list(self.proteinSeq)[j])]
					filewriter.writerow(lines)
					lines=[]
					i=i+1
				i=0
				j=j+1
			csvoutfile.close()	
		elif(self.Ensembl==False):
			k=k
			csvoutfile=open((self.output_file), 'w',newline="") 	
			filewriter=csv.writer(csvoutfile, delimiter='\t' )
			filewriter.writerow(np.array(self.headerValues))
			j=0
			i=0
			while( j< len(self.accessionId)):
				while (i< len(list(self.mutSYSeq)[j]) and i<len(list(self.regSYSeq)[j])):
					lines=[str(list(self.accessionId)[j]), str(list(self.mutIndex)[j]), str(list(self.aChange)[j]), str(list(self.mutSeq)[j]),
					str(list(self.regSeq)[j]),
					str(list(self.mutSYSeq)[j][i]),
					str(list(self.regSYSeq)[j][i]),
					str(list(self.mutSYStrength)[j][i]),
					str(list(self.regSYStrength)[j][i]),
					str(list(self.proteinSeq)[j])]
					filewriter.writerow(lines)
					lines=[]
					i=i+1
				i=0
				j=j+1
			csvoutfile.close()
		return
		
	
	
	def writeToXSLX(self, i):
		j=i+1

		self.mutSeqWrite(j)
		self.regSeqWrite(j)
		self.mutSYSeqWrite(j)
		self.mutSYStrengthWrite( j)	
		self.regSYStrengthWrite( j)	
		self.proteinSeqWrite(j)
		self.regSYSeqWrite(j)
		return
	def accessionIdWrite(self, j):
		i=j
		
		self.colWrite(self.accessionId, i, self.accessionId_col)
	def mutIndexWrite(self, j):
		i=j
		self.colWrite(self.mutIndex, i, self.mutIndex_col)
	def bChangeWrite(self,j):
		i=j
		self.colWrite(self.aChange, i, self.aChange_col)
	def mutSeqWrite(self,j):
		i=j
		self.colWrite(self.mutSeq, i, self.mutSeq_col)
	def regSeqWrite(self,j):
		i=j
		self.colWrite(self.regSeq, i, self.regSeq_col)	
	def mutSYSeqWrite(self,j):
		i=j
		self.colWrite(list(self.mutSYSeq), i, self.mutSYSeq_col)
	def mutSYStrengthWrite(self, j):
		i=j
		self.colWrite(list(self.mutSYStrength), i, self.mutSYStrength_col)	
	def regSYStrengthWrite(self, j):
		i=j
		self.colWrite(self.regSYStrength, i, self.regSYStrength_col)
	def proteinSeqWrite(self, j):
		i=j
		self.colWrite( self.proteinSeq, i, self.proteinSeq_col) 
	def colWrite(self, fullList, num, col):
		colum=str(col)
		fname =  str(self.input_file)
		j = int(num)+1
		sname= self.input_sheet
		wb = load_workbook(filename=fname)
		ws= wb.get_sheet_by_name(sname)
		i=0
		while (i<len(fullList)):
			inPut = str(fullList[i]) 
			realRow=i+j
			idCellNum=str(colum+str(realRow))
			ws[idCellNum].value=inPut
			i= i+1
		wb.save(fname)
		return
	def regSYSeqWrite(self,j):
		i=j
		self.colWrite(self.regSYSeq, i, self.regSYSeq_col)	
	
	def mutSYFEITHIProcessing(self):
		print("Mut SY")
		i=0
		j=0
		merLength=int(self.amerLength)
		while i< len(self.mutSeq):
			try:
				if(str((self.mutSeq)[i])!='' and len(str(self.mutSeq[i]))>=merLength and (str(self.mutSeq[i]).find('?'))<0):
					print(self.mutSeq[i])	
					self.lookUpSYFEITHI((self.mutSeq[i]))
					self.mutSYSeq.append((self.SYSeq))
					self.mutSYStrength.append((self.SYStrength))
					print("Done Mutation! with " + str(i))	
				elif(str(self.mutSeq[i])=='-' or len(str(self.mutSeq[i]))<merLength):
					self.mutSYSeq.append('-')
					self.mutSYStrength.append('-')
				i=i+1
				j=j+1
			except:
				i=i+1
		print("done")
		return 
	
	def regDuplicateSYFEITHIProcessing(self):
		merLength=self.amerLength
		amerLength=merLength-1
		i=0
		j=0
		while i< len(self.regSeq):
			try:
				if(str(self.regSeq[i])=='-' or len(str(self.regSeq[i]))<merLength or len(str(self.mutSYSeq[i]))<merLength ):
					self.regSYSeq.append('-')
					self.regSYStrength.append('-')
				elif(self.regSeq[i]!='' and len(str(self.regSeq[i]))>=merLength and len(str(self.mutSYSeq[i]))>=merLength):
					print(self.regSeq[i])	
					regSeqStart=(self.mutSeq[i]).find(str(self.mutSYSeq[i]))
					regSYSeq=(self.regSeq[i])[regSeqStart:regSeqStart+merLength]
					self.lookUpSYFEITHI(regSYSeq)
					self.regSYStrength.append(self.SYStrength)
					self.regSYSeq.append(self.SYSeq)
					time.sleep(1)
					#print("Done! with " + str(i))
			except ValueError:
				i=i+1
				j=j+1
				break
					
			i=i+1
			j=j+1
		#self.regSYSeqWrite(writeLine)
		return 		
	
	def regSYFEITHIProcessing(self):
		#self.readInputAccessionId()
		print("Reg SY")
		merLength=int(self.amerLength)
		i=0
		writeLine=0
		j=0
		try:
			while i< len(self.regSeq):
				if(str(self.regSeq[i])=='-' or len(str(self.regSeq[i]))<merLength):
					self.regSYSeq.append('-')
					self.regSYStrength.append('-')
				elif(self.regSeq[i]!='' and len(str(self.regSeq[i]))>=merLength):
					#print("starting SYPFIETHI")
					#print(self.regSeq[i])	
					self.lookUpSYFEITHI(self.regSeq[i])
					self.regSYStrength.append((self.SYStrength))
					self.regSYSeq.append((self.SYSeq))
					#print("Done! with " + str(i))	
				i=i+1
				j=j+1
		except:
				i=i+1
				j=j+1
		print("done")
		return 		

	def lookUpNetCTLPan(self, sequ):
		
		browse= Browser()
	
		Query="http://www.cbs.dtu.dk/services/NetCTLpan/"
		net_page = browse.get(Query)
		net_form= browse.select("")
		return

	def lookUpSYFEITHI(self,sequ):
		seqN=sequ
		while True:
			try:
				Query="http://www.syfpeithi.de/bin/MHCServer.dll/EpitopePrediction?Motif=" + str(self.geneType)+ "&amers="+ str(self.amerLength) + "&SEQU=" + str(seqN)+"&DoIT=++Run++"
				response = urllib.request.Request(Query)
				html = urllib.request.urlopen(response)
				readText=str(html.read())
				#print(readText)
				html.close()
						#1 first need to isolate the table 
				start=int(readText.find('<TR'))
				end=int(readText.find('</tr></table>'))
				readText=readText[start:end]
				#print(readText)
				
				readText=readText.replace('&nbsp;','')
				readText=readText.replace('</td>','')
				readText=readText.replace('<U>', '')
				readText=readText.replace('<B>', '')
				readText=readText.replace('</U>', '')
				readText=readText.replace('</B>', '')
				readText=readText.replace('<TT>','')
				#print(readText)
				readText=list(readText.split('</tr>'))
				#print(readText)
				i=0
				syseq=[]
				systrength=[]
				while (i<len(readText) and i<int(self.numTopSY)):
					readText[i]=readText[i].replace('<td align=right>',',')
					readText[i]=readText[i].replace('<td nowrap align=center>',',')
					readText[i]=readText[i].split(',')
					readText[i]=readText[i][2:]
					syseq.append(readText[i][0])
					systrength.append(readText[i][1])
					#print(readText[i])	
					i=i+1
				self.SYSeq=syseq
				self.SYStrength=systrength
				return 
			except RuntimeError:
				time.sleep(5) 
	def processMutationLists(self):
		if(self.Ensembl):
			self.readInputEnsembl()
			i=0
			k=-1
			writeLine=1
			try:
				while (i<len((self.EnsemblId))):
					if( str(self.aChange[i])!='-' and str(self.aChange[i])!=''):
						#print("Found One!" + str(i))
						if(i<1):
							self.lookUpEnsemblProtein(i)
						elif(self.EnsemblId[i]!= self.EnsemblId[k]):
							self.lookUpEnsemblProtein(i)
						elif(self.EnsemblId[i]== self.EnsemblId[k]):
							protein=self.proteinSeq[k]
							self.proteinSeq.append(protein)
					elif( str(self.aChange[i])=='-' or str(self.aChange[i])==' '):
						self.proteinSeq.append('-')
					i=i+1
					k=k+1
				print("done with proteins")
				if(str(self.amerLength)=='0'): 
					for i in self.amerLengthList:
						self.amerLength=i	
						self.getMutSeq()
						self.mutSYFEITHIProcessing()
						self.getRegSeq()
						self.regSYFEITHIProcessing()
				elif(str(self.amerLength)!='0'):	
						#print("getting mutations")
						self.getRegSeq()
						self.getMutSeq()
						self.regSYFEITHIProcessing()
						self.mutSYFEITHIProcessing()
						
				return
			except:
				if(self.amerLength==0): 
					for i in self.amerLengthList:
						self.amerLength=i	
						self.getMutSeq()
						self.mutSYFEITHIProcessing()
						self.getRegSeq()
						self.regSYFEITHIProcessing()
				else:
					self.getMutSeq()
					self.getRegSeq()
					self.regSYFEITHIProcessing()
					self.mutSYFEITHIProcessing()
				return
	
		else:
			self.readInputAccessionId()
			i=0
			j=0
			k=-1
			writeLine=1
			try:
				
				while (self.accessionId[i]!=' ') and (self.accessionId[i])!='None':
					if( self.aChange[i]!='-' and self.aChange[i]!=''):
						#print("Found One!" + str(i))
						if(i<1):
							self.lookUpProtein(i)
						elif(self.accessionId[i]!= self.accessionId[k]):
							self.lookUpProtein(i)
						elif(self.accessionId[i]== self.accessionId[k]):
							protein=self.proteinSeq[k]
							self.proteinSeq.append(protein)
					elif( self.aChange[i]=='-' or self.aChange[i]==' '):
						self.proteinSeq.append('-')
					i=i+1
					k=k+1
				print("done with proteins")		   	
				self.getMutSeq()
				self.getRegSeq()
				print("got mutations")
				self.mutSYFEITHIProcessing()
				self.regSYFEITHIProcessing()
				print("Done with Processing")
			except:
				self.getMutSeq()
				self.getRegSeq()
				self.regSYFEITHIProcessing()
				self.mutSYFEITHIProcessing()
				print("Done with Processing")
				return
				
	def __init__(self, input_file=None,input_sheet=None, output_file=None, output_sheet=None, input_list=None, amerLength='9', geneType="HLA-A*02%3A01", numSY=1, Ensembl=False):
		self.input_file=str(input_file)
		self.input_sheet=str(input_sheet)
		self.input_list=(input_list)
		self.output_file=str(output_file)
		self.output_sheet=str(output_sheet)
		self.amerLength=amerLength
		self.amerLengthList=[8,9,10,11,15,0]
		self.geneType=(geneType)
		self.numTopSY=numSY
		self.accessionId=[]
		self.mutIndex=[]
		self.Ensembl=Ensembl
		self.isText=False
		self.aChange=[]
		self.mutSeq=[]
		self.regSeq=[]
		self.mutSYSeq=[]
		self.regSYSeq=[]
		self.mutSYStrength=[]
		self.regSYStrength=[]
		self.proteinSeq=[]
		self.EnsemblId=[]
		self.SYSeq=[]
		self.SYStrength=[]
		self.fullEnsembl=[]
		self.accessionId_col=str(self.input_list[0])
		self.Ensembl_col= str(self.input_list[1])
		self.mutIndex_col=str(self.input_list[2])
		self.aChange_col=str(self.input_list[3])
		self.mutSeq_col=str(self.input_list[4])
		self.regSeq_col=str(self.input_list[5])
		self.mutSYSeq_col=str(self.input_list[6])
		self.regSYSeq_col=str(self.input_list[7])
		self.mutSYStrength_col=str(self.input_list[8])
		self.regSYStrength_col=str(self.input_list[9])
		self.proteinSeq_col=str(self.input_list[10])
		
		self.numDict={'A': 0, 'B': 1, 'C':2, 'D': 3, 'E': 4, 'F' : 5, 'G':6 ,'H' :7, 'I' :8, 'J' :9, 'K' : 10, 'L' : 11, 'M': 12, 'N' : 13,  'O': 14,   'P': 15,  'Q': 16, 'R':17, 'S':18, 'T':19, 'U':20,'V':21, 'W':22, 'X':23, 'Y':24 ,'Z':25, 'AA':26, 'AB':27, 'AC': 28, 'AD':29, 'AE':30 ,'AF': 31} 
		self.aminoAcidDict={ 'Gly' : 'G', 'Pro' : 'P' ,'Ala' :'A', 'Val' :'V' , 'Leu' :'L', 'Ile':'I',  'Met' :'M', 'Cys': 'C','Phe': 'F', 'Tyr': 'Y', 'Trp' :'W', 'His' :'H', 'Lys' :'K','Arg' :'R','Gln':'Q', 'Asn' :'N','Glu' :'E','Asp': 'D','Ser':'S','Thr' :'T', '*':'*', '?':'?'}
		self.headerValues=['Ensembl Id','Mutation Index', 'Amino Acid Change', 'Mutation Sequence','Regular Sequence','Mutation SYPFIETHI Binding','Regular SYPFIETHI Binding','Mutation SYPFIETHI Strength','Regular SYPFEITHI Strength','Protein']

		
		

excel=Excel()
excel.lookUpNetCTLPan("HEHFHDBSHFHDHSJ")


		